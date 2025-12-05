import openpyxl
from openpyxl.styles import PatternFill
import os

class ExcelUtils:
    def __init__(self, file_path, sheet_name=None):
        # Resolve absolute path
        base_dir = os.path.dirname(os.path.abspath(__file__))
        resolved_path = os.path.join(base_dir, '..', file_path)
        self.file_path = os.path.normpath(resolved_path)

        self.wb = openpyxl.load_workbook(self.file_path)

        # Use provided sheet name or default to 'Sheet1'
        self.sheet_name = sheet_name if sheet_name else "Sheet1"
        self.sheet = self.wb[self.sheet_name]
        self.header = [cell.value for cell in self.sheet[1]]

    def get_test_data(self):
        data = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_data = dict(zip(self.header, row))
            data.append(row_data)
        return data

    @staticmethod
    def read_test_data(file_path, sheet_name="Sheet1"):
        excel_util = ExcelUtils(file_path, sheet_name)
        return excel_util.get_test_data()

    def write_result(self, tc_name, status):
        color_map = {
            "PASS": "00FF00",
            "FAIL": "FF0000",
            "SKIP": "FFFF00",
        }
        for i, row in enumerate(self.sheet.iter_rows(min_row=2), start=2):
            if row[0].value == tc_name:
                result_cell = self.sheet.cell(row=i, column=len(self.header) + 1)
                result_cell.value = status
                result_cell.fill = PatternFill(start_color=color_map[status], fill_type="solid")
                break
        self.wb.save(self.file_path)
